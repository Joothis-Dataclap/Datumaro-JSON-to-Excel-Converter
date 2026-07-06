import streamlit as st
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import time


@st.cache_data
def parse_datumaro_json(json_data):
    """
    Parse Datumaro JSON format and extract annotation data
    Returns: list of dicts with image_name, frame_number, and label annotations
    """
    try:
        data = json.loads(json_data) if isinstance(json_data, str) else json_data
        
        # Build label_id to label_name mapping from categories
        label_map = {}
        if "categories" in data and "label" in data["categories"]:
            labels = data["categories"]["label"].get("labels", [])
            for idx, label_obj in enumerate(labels):
                label_map[idx] = label_obj.get("name", f"Label_{idx}")
        
        # Extract items from Datumaro format
        items = data.get("items", [])
        annotations_data = []
        
        for frame_number, item in enumerate(items, start=0):
            image_name = item.get("id", "")
            
            # Extract annotations for this image
            annotations = item.get("annotations", [])
            label_counts = {}
            
            for annotation in annotations:
                label_id = annotation.get("label_id", -1)
                label_name = label_map.get(label_id, f"Unknown_{label_id}")
                annotation_type = annotation.get("type", "")
                
                if label_name not in label_counts:
                    label_counts[label_name] = {"box": 0, "polygon": 0}
                
                # Count by type
                if annotation_type == "bbox":
                    label_counts[label_name]["box"] += 1
                elif annotation_type == "polygon":
                    label_counts[label_name]["polygon"] += 1
            
            annotations_data.append({
                "image_name": image_name,
                "frame_number": frame_number,
                "label_counts": label_counts
            })
        
        return annotations_data
    
    except json.JSONDecodeError as e:
        st.error(f"Invalid JSON format: {str(e)}")
        return None
    except Exception as e:
        st.error(f"Error parsing JSON: {str(e)}")
        return None


@st.cache_data
def create_excel_with_labels(json_data):
    """
    Create Excel file with dynamic label columns
    Structure: SI No | Image Name | Frame Number | [Label 1 Box | Label 1 Polygon] | ... | Total Annotations
    """
    annotations_data = parse_datumaro_json(json_data)
    if not annotations_data:
        return None, None
    
    # Collect all unique labels
    all_labels = set()
    for item in annotations_data:
        all_labels.update(item["label_counts"].keys())
    all_labels = sorted(list(all_labels))
    
    # Prepare data for DataFrame
    rows = []
    for si, item in enumerate(annotations_data, 1):
        row = {
            "SI No": si,
            "Image Name": item["image_name"],
            "Frame Number": item["frame_number"]
        }
        
        total_annotations = 0
        for label in all_labels:
            box_count = item["label_counts"].get(label, {}).get("box", 0)
            polygon_count = item["label_counts"].get(label, {}).get("polygon", 0)
            
            row[f"{label} - Box"] = box_count
            row[f"{label} - Polygon"] = polygon_count
            total_annotations += box_count + polygon_count
        
        row["Total Annotations"] = total_annotations
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"
    
    # Write headers
    headers = ["SI No", "Image Name", "Frame Number"]
    for label in all_labels:
        headers.append(f"{label} - Box")
        headers.append(f"{label} - Polygon")
    headers.append("Total Annotations")
    
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data rows
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_data in rows:
        row_values = [
            row_data["SI No"],
            row_data["Image Name"],
            row_data["Frame Number"]
        ]
        for label in all_labels:
            row_values.append(row_data[f"{label} - Box"])
            row_values.append(row_data[f"{label} - Polygon"])
        row_values.append(row_data["Total Annotations"])
        
        ws.append(row_values)
    
    # Apply borders and alignment to all cells
    for row in ws.iter_rows(min_row=2, max_row=len(rows)+1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Add summary row with totals
    summary_row = len(rows) + 3
    ws[f'A{summary_row}'] = "SUMMARY - Total Counts:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    # Add summary formulas for ALL columns (from D onwards to last column)
    for col_idx in range(4, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{summary_row}']
        cell.value = f'=SUM({col_letter}2:{col_letter}{len(rows)+1})'
        cell.font = Font(bold=True)
        
        # Color coding: Light green for label columns, light red for total
        if col_idx == len(headers):  # Total Annotations column
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    return wb, df


@st.cache_data
def parse_polygon_points_data(json_data):
    """
    Parse Datumaro JSON and extract per-polygon point data.
    Returns list of dicts: image_name, label_name, polygon_id, points_count
    """
    try:
        data = json.loads(json_data) if isinstance(json_data, str) else json_data

        label_map = {}
        if "categories" in data and "label" in data["categories"]:
            labels = data["categories"]["label"].get("labels", [])
            for idx, label_obj in enumerate(labels):
                label_map[idx] = label_obj.get("name", f"Label_{idx}")

        items = data.get("items", [])
        polygon_rows = []

        for item in items:
            image_name = item.get("id", "")
            annotations = item.get("annotations", [])

            for annotation in annotations:
                if annotation.get("type", "") == "polygon":
                    label_id = annotation.get("label_id", -1)
                    label_name = label_map.get(label_id, f"Unknown_{label_id}")
                    polygon_id = annotation.get("id", "N/A")
                    points = annotation.get("points", [])
                    points_count = len(points) // 2

                    polygon_rows.append({
                        "image_name": image_name,
                        "label_name": label_name,
                        "polygon_id": polygon_id,
                        "points_count": points_count
                    })

        return polygon_rows

    except json.JSONDecodeError as e:
        st.error(f"Invalid JSON format: {str(e)}")
        return None
    except Exception as e:
        st.error(f"Error parsing JSON: {str(e)}")
        return None


def annotation_converter_page():
    st.title("🔄 Datumaro JSON to Excel Converter")

    uploaded_file = st.file_uploader(
        "Choose a Datumaro JSON file",
        type="json",
        key="converter_uploader"
    )

    if uploaded_file is not None:
        try:
            json_content = uploaded_file.read().decode("utf-8")

            msg_placeholder = st.empty()
            msg_placeholder.info("📋 Processing JSON file...")

            wb, df = create_excel_with_labels(json_content)

            if df is None or df.empty:
                st.error("No valid annotation data found in JSON file")
                return

            msg_placeholder.success(f"✅ Found {len(df)} images with annotations")
            time.sleep(3)
            msg_placeholder.empty()

            df['Job'] = df['Image Name'].str.split('_').str[0]

            label_cols = [col for col in df.columns if "- Box" in col]
            selected_labels = sorted([col.replace(" - Box", "") for col in label_cols])

            filtered_df = df.copy()

            st.divider()

            col_stats, col_download = st.columns([2, 1])

            with col_stats:
                st.subheader("📈 Statistics")
                stat_col1, stat_col2, stat_col3 = st.columns(3)

                with stat_col1:
                    st.metric("Total Images", len(filtered_df))

                with stat_col2:
                    st.metric("Total Annotations", int(filtered_df["Total Annotations"].sum()))

                with stat_col3:
                    st.metric("Unique Labels", len(selected_labels))

            with col_download:
                st.subheader("💾 Download")

                output_buffer = io.BytesIO()

                filtered_df_for_export = filtered_df.copy()
                wb_filtered = Workbook()
                ws = wb_filtered.active
                ws.title = "Annotations"

                headers = ["SI No", "Job", "Image Name", "Frame Number"]
                for label in selected_labels:
                    headers.append(f"{label} - Box")
                    headers.append(f"{label} - Polygon")
                headers.append("Total Annotations of the Image")

                ws.append(headers)

                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for idx, row_data in filtered_df_for_export.iterrows():
                    row_values = [
                        row_data["SI No"],
                        row_data["Job"],
                        row_data["Image Name"],
                        row_data["Frame Number"]
                    ]
                    for label in selected_labels:
                        row_values.append(row_data[f"{label} - Box"])
                        row_values.append(row_data[f"{label} - Polygon"])
                    row_values.append(row_data["Total Annotations"])

                    ws.append(row_values)

                for row in ws.iter_rows(min_row=2, max_row=len(filtered_df_for_export)+1):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 15
                for col in range(5, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 14

                summary_row = len(filtered_df_for_export) + 3
                ws[f'A{summary_row}'] = "SUMMARY - Total Counts:"
                ws[f'A{summary_row}'].font = Font(bold=True)

                for col_idx in range(5, len(headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    cell = ws[f'{col_letter}{summary_row}']
                    cell.value = f'=SUM({col_letter}2:{col_letter}{len(filtered_df_for_export)+1})'
                    cell.font = Font(bold=True)

                    if col_idx == len(headers):
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

                wb_filtered.save(output_buffer)
                output_buffer.seek(0)

                filename = f"{uploaded_file.name.split('.')[0]}_annotations_filtered.xlsx"

                st.download_button(
                    label="📥 Download Excel",
                    data=output_buffer.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            st.divider()

            st.subheader("📊 Data Preview")
            col_order = ["SI No", "Job", "Image Name", "Frame Number"]
            for label in selected_labels:
                col_order.append(f"{label} - Box")
                col_order.append(f"{label} - Polygon")
            col_order.append("Total Annotations")

            display_df = filtered_df[col_order]
            st.dataframe(display_df, use_container_width=True, height=500, hide_index=True)

            st.divider()
            st.subheader("🏷️ Annotation Breakdown by Label")

            if selected_labels:
                label_stats = {}
                for label in selected_labels:
                    box_col = f"{label} - Box"
                    polygon_col = f"{label} - Polygon"

                    box_count = filtered_df[box_col].sum()
                    polygon_count = filtered_df[polygon_col].sum()

                    label_stats[label] = {
                        "Box": int(box_count),
                        "Polygon": int(polygon_count),
                        "Total": int(box_count + polygon_count)
                    }

                stats_df = pd.DataFrame(label_stats).T
                st.dataframe(stats_df, use_container_width=True)

        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            import traceback
            st.error(traceback.format_exc())


def polygon_point_counter_page():
    st.title("🔷 Polygon Point Counter")
    st.markdown("Analyze per-polygon vertex counts from Datumaro JSON files.")

    uploaded_file = st.file_uploader(
        "Choose a Datumaro JSON file",
        type="json",
        key="polygon_uploader"
    )

    if uploaded_file is not None:
        try:
            json_content = uploaded_file.read().decode("utf-8")

            msg_placeholder = st.empty()
            msg_placeholder.info("🔷 Extracting polygon point data...")

            polygon_rows = parse_polygon_points_data(json_content)

            if not polygon_rows:
                st.error("No polygon annotations found in the JSON file.")
                return

            msg_placeholder.success(f"✅ Found {len(polygon_rows)} polygons")
            time.sleep(2)
            msg_placeholder.empty()

            # Extract unique labels for checkbox selection
            all_labels = sorted(list(set(row["label_name"] for row in polygon_rows)))

            # Add checkbox section for label selection
            st.subheader("🏷️ Select Labels to Include")
            selected_labels = st.multiselect(
                "Choose which labels to include in the analysis and download:",
                options=all_labels,
                default=all_labels,
                key="label_filter"
            )

            # Validation: show warning if no labels selected
            if not selected_labels:
                st.warning("⚠️ Please select at least one label to proceed.")
                return

            # Filter polygon_rows based on selected labels
            filtered_polygon_rows = [row for row in polygon_rows if row["label_name"] in selected_labels]

            df = pd.DataFrame(filtered_polygon_rows)[["image_name", "label_name", "points_count"]].copy()
            df.insert(0, "SI No", range(1, len(df) + 1))
            df.columns = ["SI No", "Image Name", "Label Name", "Points Count"]

            st.divider()

            col_stats, col_download = st.columns([2, 1])

            with col_stats:
                st.subheader("📈 Statistics")
                s1, s2 = st.columns(2)
                with s1:
                    st.metric("Total Polygons", len(df))
                with s2:
                    st.metric("Images with Polygons", df["Image Name"].nunique())

            with col_download:
                st.subheader("💾 Download")

                wb_poly = Workbook()
                ws = wb_poly.active
                ws.title = "Polygon Points"

                poly_headers = ["SI No", "Image Name", "Label Name", "Points Count"]
                ws.append(poly_headers)

                header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for _, row_data in df.iterrows():
                    ws.append([
                        int(row_data["SI No"]),
                        row_data["Image Name"],
                        row_data["Label Name"],
                        int(row_data["Points Count"])
                    ])

                for row in ws.iter_rows(min_row=2, max_row=len(df) + 1):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=2, max_col=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 15

                summary_row = len(df) + 3
                ws[f'A{summary_row}'] = "TOTAL POLYGONS:"
                ws[f'A{summary_row}'].font = Font(bold=True)
                poly_total_cell = ws[f'C{summary_row}']
                poly_total_cell.value = len(df)
                poly_total_cell.font = Font(bold=True)
                poly_total_cell.fill = PatternFill(start_color="E2D0F0", end_color="E2D0F0", fill_type="solid")

                ws[f'A{summary_row + 1}'] = "TOTAL POINTS:"
                ws[f'A{summary_row + 1}'].font = Font(bold=True)
                pts_total_cell = ws[f'D{summary_row + 1}']
                pts_total_cell.value = f'=SUM(D2:D{len(df) + 1})'
                pts_total_cell.font = Font(bold=True)
                pts_total_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                ws2 = wb_poly.create_sheet("Label Summary")
                sum_headers = ["Label Name", "Polygon Count", "Total Points"]
                ws2.append(sum_headers)

                for cell in ws2[1]:
                    cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                label_groups = {}
                for row in filtered_polygon_rows:
                    ln = row["label_name"]
                    if ln not in label_groups:
                        label_groups[ln] = {"count": 0, "total_points": 0}
                    label_groups[ln]["count"] += 1
                    label_groups[ln]["total_points"] += row["points_count"]

                for label, stats in sorted(label_groups.items()):
                    ws2.append([label, stats["count"], stats["total_points"]])

                for row in ws2.iter_rows(min_row=2, max_row=len(label_groups) + 1):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 15
                ws2.column_dimensions['C'].width = 15

                output_buffer = io.BytesIO()
                wb_poly.save(output_buffer)
                output_buffer.seek(0)

                filename = f"{uploaded_file.name.split('.')[0]}_polygon_points.xlsx"

                st.download_button(
                    label="📥 Download Excel",
                    data=output_buffer.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            st.divider()

            st.subheader("🔷 Polygon Points Data")
            st.dataframe(df, use_container_width=True, height=500, hide_index=True)

            st.divider()
            st.subheader("📊 Per-Label Summary")

            label_summary_df = df.groupby("Label Name").agg(
                Polygon_Count=("Points Count", "count"),
                Total_Points=("Points Count", "sum")
            ).reset_index()
            label_summary_df.columns = ["Label Name", "Polygon Count", "Total Points"]
            st.dataframe(label_summary_df, use_container_width=True, hide_index=True)

            st.divider()
            total_col1, total_col2 = st.columns(2)
            with total_col1:
                st.metric("🔢 Total Polygons", len(df))
            with total_col2:
                st.metric("📍 Total Points (All Polygons)", int(df["Points Count"].sum()))

        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            import traceback
            st.error(traceback.format_exc())


def filter_json_by_labels(json_data, selected_labels, keep_selected=True):
    """
    Filter Datumaro JSON to keep only annotations for selected or non-selected labels.
    Args:
        json_data: Datumaro JSON as string or dict
        selected_labels: List of label names to filter by
        keep_selected: If True, keep selected labels; if False, keep non-selected labels
    Returns filtered JSON as string.
    """
    try:
        data = json.loads(json_data) if isinstance(json_data, str) else json_data

        # Get label_id to label_name mapping
        label_map = {}
        if "categories" in data and "label" in data["categories"]:
            labels = data["categories"]["label"].get("labels", [])
            for idx, label_obj in enumerate(labels):
                label_map[idx] = label_obj.get("name", f"Label_{idx}")

        # Create reverse mapping: label_name to label_id
        name_to_id = {v: k for k, v in label_map.items()}

        # Filter items: keep only annotations for selected or non-selected labels
        filtered_items = []
        for item in data.get("items", []):
            filtered_annotations = []
            for annotation in item.get("annotations", []):
                label_id = annotation.get("label_id", -1)
                label_name = label_map.get(label_id, f"Unknown_{label_id}")
                if keep_selected:
                    # Keep selected labels
                    if label_name in selected_labels:
                        filtered_annotations.append(annotation)
                else:
                    # Keep non-selected labels
                    if label_name not in selected_labels:
                        filtered_annotations.append(annotation)

            # Only keep item if it has annotations after filtering
            if filtered_annotations:
                item_copy = item.copy()
                item_copy["annotations"] = filtered_annotations
                filtered_items.append(item_copy)

        # Update categories to include only selected or non-selected labels
        if "categories" in data and "label" in data["categories"]:
            filtered_labels = []
            new_label_map = {}
            new_id = 0
            for label_obj in data["categories"]["label"].get("labels", []):
                label_name = label_obj.get("name", "")
                if keep_selected:
                    # Keep selected labels
                    if label_name in selected_labels:
                        new_label_map[new_id] = label_name
                        label_obj_copy = label_obj.copy()
                        label_obj_copy["id"] = new_id
                        filtered_labels.append(label_obj_copy)
                        new_id += 1
                else:
                    # Keep non-selected labels
                    if label_name not in selected_labels:
                        new_label_map[new_id] = label_name
                        label_obj_copy = label_obj.copy()
                        label_obj_copy["id"] = new_id
                        filtered_labels.append(label_obj_copy)
                        new_id += 1

            data["categories"]["label"]["labels"] = filtered_labels

            # Update label_ids in annotations to match new mapping
            old_to_new_id = {}
            for old_id, name in label_map.items():
                if keep_selected:
                    if name in selected_labels:
                        for new_id, new_name in new_label_map.items():
                            if new_name == name:
                                old_to_new_id[old_id] = new_id
                                break
                else:
                    if name not in selected_labels:
                        for new_id, new_name in new_label_map.items():
                            if new_name == name:
                                old_to_new_id[old_id] = new_id
                                break

            for item in filtered_items:
                for annotation in item["annotations"]:
                    old_label_id = annotation.get("label_id", -1)
                    if old_label_id in old_to_new_id:
                        annotation["label_id"] = old_to_new_id[old_label_id]

        data["items"] = filtered_items
        return json.dumps(data, indent=2)

    except Exception as e:
        raise Exception(f"Error filtering JSON: {str(e)}")


def merge_datumaro_jsons(json_files_list):
    """
    Merge multiple Datumaro JSON files into one.
    Returns merged JSON as string.
    Raises Exception if duplicate label names are found across files.
    """
    try:
        all_data = []
        for json_content in json_files_list:
            data = json.loads(json_content) if isinstance(json_content, str) else json_content
            all_data.append(data)

        # Check for duplicate label names across all files
        all_labels_with_file = []
        for idx, data in enumerate(all_data):
            if "categories" in data and "label" in data["categories"]:
                for label_obj in data["categories"]["label"].get("labels", []):
                    label_name = label_obj.get("name", "")
                    all_labels_with_file.append((label_name, idx))

        # Find duplicates
        label_names = [label[0] for label in all_labels_with_file]
        unique_labels = set(label_names)
        if len(label_names) != len(unique_labels):
            # Find which labels are duplicated
            from collections import Counter
            label_counts = Counter(label_names)
            duplicates = {label: count for label, count in label_counts.items() if count > 1}

            # Build error message with file information
            error_msg = "Duplicate label names found across files:\n"
            for dup_label in duplicates:
                files_with_label = [f"File {idx+1}" for label, idx in all_labels_with_file if label == dup_label]
                error_msg += f"- '{dup_label}' appears in: {', '.join(files_with_label)}\n"
            error_msg += "\nPlease remove or rename duplicate labels before merging."
            raise Exception(error_msg)

        # Merge categories: combine all unique labels
        merged_labels = []
        label_name_to_new_id = {}
        new_id = 0

        for data in all_data:
            if "categories" in data and "label" in data["categories"]:
                for label_obj in data["categories"]["label"].get("labels", []):
                    label_name = label_obj.get("name", f"Label_{new_id}")
                    if label_name not in label_name_to_new_id:
                        label_name_to_new_id[label_name] = new_id
                        merged_labels.append({
                            "id": new_id,
                            "name": label_name,
                            **{k: v for k, v in label_obj.items() if k not in ["id", "name"]}
                        })
                        new_id += 1

        # Build old label_id to new label_id mapping for each file
        file_mappings = []
        for data in all_data:
            old_to_new = {}
            if "categories" in data and "label" in data["categories"]:
                for label_obj in data["categories"]["label"].get("labels", []):
                    old_id = label_obj.get("id", -1)
                    label_name = label_obj.get("name", "")
                    if label_name in label_name_to_new_id:
                        old_to_new[old_id] = label_name_to_new_id[label_name]
            file_mappings.append(old_to_new)

        # Merge items: combine all items, merging annotations for same image_id
        merged_items = {}
        for idx, data in enumerate(all_data):
            mapping = file_mappings[idx]
            for item in data.get("items", []):
                image_id = item.get("id", "")
                
                if image_id not in merged_items:
                    merged_items[image_id] = {
                        "id": image_id,
                        "annotations": []
                    }
                
                # Add annotations with updated label_ids
                for annotation in item.get("annotations", []):
                    annotation_copy = annotation.copy()
                    old_label_id = annotation_copy.get("label_id", -1)
                    if old_label_id in mapping:
                        annotation_copy["label_id"] = mapping[old_label_id]
                    merged_items[image_id]["annotations"].append(annotation_copy)

        # Build merged JSON structure
        merged_data = {
            "categories": {
                "label": {
                    "labels": merged_labels
                }
            },
            "items": list(merged_items.values())
        }

        return json.dumps(merged_data, indent=2)

    except Exception as e:
        raise Exception(f"Error merging JSONs: {str(e)}")


def json_split_merge_page():
    st.title("🔀 JSON Split & Merge")
    st.markdown("Filter annotations by label or merge multiple JSON files.")

    # Mode selection
    mode = st.radio(
        "Select Mode:",
        ["🔍 Split/Filter by Labels", "🔗 Merge Multiple JSONs"],
        label_visibility="collapsed",
        horizontal=True
    )

    st.divider()

    if mode == "🔍 Split/Filter by Labels":
        st.subheader("🔍 Filter JSON by Labels")
        st.markdown("Upload a Datumaro JSON file and select which labels to keep.")

        uploaded_file = st.file_uploader(
            "Choose a Datumaro JSON file",
            type="json",
            key="filter_uploader"
        )

        if uploaded_file is not None:
            try:
                json_content = uploaded_file.read().decode("utf-8")
                data = json.loads(json_content)

                # Extract all unique labels
                all_labels = []
                if "categories" in data and "label" in data["categories"]:
                    for label_obj in data["categories"]["label"].get("labels", []):
                        all_labels.append(label_obj.get("name", f"Unknown"))
                all_labels = sorted(list(set(all_labels)))

                if not all_labels:
                    st.error("No labels found in the JSON file.")
                    return

                st.info(f"Found {len(all_labels)} labels in the file.")

                # Label selection
                st.subheader("🏷️ Select Labels to Keep")
                selected_labels = st.multiselect(
                    "Choose which labels to include in the filtered JSON:",
                    options=all_labels,
                    default=all_labels,
                    key="filter_labels"
                )

                if not selected_labels:
                    st.warning("⚠️ Please select at least one label to proceed.")
                    return

                # Filter and preview
                if st.button("🔄 Filter JSON", use_container_width=True):
                    with st.spinner("Filtering JSON..."):
                        # Filter selected labels
                        filtered_json_selected = filter_json_by_labels(json_content, selected_labels, keep_selected=True)
                        filtered_data_selected = json.loads(filtered_json_selected)

                        # Filter non-selected labels
                        other_labels = [label for label in all_labels if label not in selected_labels]
                        filtered_json_other = filter_json_by_labels(json_content, selected_labels, keep_selected=False)
                        filtered_data_other = json.loads(filtered_json_other)

                        st.success(f"✅ Filtered JSONs created")

                        # Show statistics side by side with vertical divider
                        col_stats_left, col_divider, col_stats_right = st.columns([1, 0.05, 1])

                        with col_stats_left:
                            st.subheader("📊 Filtered JSON Statistics")
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Total Items", len(filtered_data_selected["items"]))
                            with col2:
                                total_annotations = sum(len(item.get("annotations", [])) for item in filtered_data_selected["items"])
                                st.metric("Total Annotations", total_annotations)
                            with col3:
                                st.metric("Labels Kept", len(selected_labels))

                        with col_divider:
                            st.markdown("""
                                <style>
                                    .vertical-line {
                                        border-left: 2px solid #e0e0e0;
                                        height: 100%;
                                        margin: 0 auto;
                                    }
                                </style>
                                <div class="vertical-line"></div>
                            """, unsafe_allow_html=True)

                        if other_labels:
                            with col_stats_right:
                                st.subheader("📊 Default JSON Statistics")
                                col4, col5, col6 = st.columns(3)
                                with col4:
                                    st.metric("Total Items", len(filtered_data_other["items"]))
                                with col5:
                                    total_annotations_other = sum(len(item.get("annotations", [])) for item in filtered_data_other["items"])
                                    st.metric("Total Annotations", total_annotations_other)
                                with col6:
                                    st.metric("Labels Kept", len(other_labels))

                        st.divider()

                        # Download buttons
                        st.subheader("💾 Download Filtered JSONs")
                        col_dl1, col_dl2 = st.columns(2)

                        with col_dl1:
                            filename_filtered = f"{uploaded_file.name.split('.')[0]}_filtered.json"
                            st.download_button(
                                label="📥 Download Filtered JSON",
                                data=filtered_json_selected,
                                file_name=filename_filtered,
                                mime="application/json",
                                use_container_width=True
                            )

                        if other_labels:
                            with col_dl2:
                                filename_default = f"{uploaded_file.name.split('.')[0]}_default.json"
                                st.download_button(
                                    label="📥 Download Default JSON",
                                    data=filtered_json_other,
                                    file_name=filename_default,
                                    mime="application/json",
                                    use_container_width=True
                                )

            except Exception as e:
                st.error(f"❌ Error processing file: {str(e)}")
                import traceback
                st.error(traceback.format_exc())

    else:  # Merge mode
        st.subheader("🔗 Merge Multiple JSON Files")
        st.markdown("Upload multiple Datumaro JSON files to merge them into one.")

        uploaded_files = st.file_uploader(
            "Choose Datumaro JSON files (2 or more)",
            type="json",
            accept_multiple_files=True,
            key="merge_uploader"
        )

        if uploaded_files is not None:
            if len(uploaded_files) < 2:
                st.warning("⚠️ Please upload at least 2 JSON files to merge.")
                return

            st.info(f"📁 {len(uploaded_files)} files selected for merging.")

            # Show file summary
            st.subheader("📋 Files to Merge")
            for idx, file in enumerate(uploaded_files, 1):
                st.write(f"{idx}. {file.name}")

            if st.button("🔗 Merge JSONs", use_container_width=True):
                try:
                    with st.spinner("Merging JSON files..."):
                        json_contents = []
                        for file in uploaded_files:
                            content = file.read().decode("utf-8")
                            json_contents.append(content)

                        merged_json = merge_datumaro_jsons(json_contents)
                        merged_data = json.loads(merged_json)

                        st.success(f"✅ Successfully merged {len(uploaded_files)} JSON files")

                        # Show statistics
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Items", len(merged_data["items"]))
                        with col2:
                            total_annotations = sum(len(item.get("annotations", [])) for item in merged_data["items"])
                            st.metric("Total Annotations", total_annotations)
                        with col3:
                            label_count = len(merged_data["categories"]["label"]["labels"])
                            st.metric("Total Labels", label_count)

                        st.divider()

                        # Show label summary
                        st.subheader("🏷️ Merged Labels Summary")
                        label_data = []
                        for label_obj in merged_data["categories"]["label"]["labels"]:
                            label_name = label_obj.get("name", "Unknown")
                            label_data.append({"Label Name": label_name})
                        st.dataframe(pd.DataFrame(label_data), use_container_width=True, hide_index=True)

                        st.divider()

                        # Download button
                        st.subheader("💾 Download Merged JSON")
                        st.download_button(
                            label="📥 Download Merged JSON",
                            data=merged_json,
                            file_name="merged_annotations.json",
                            mime="application/json",
                            use_container_width=True
                        )

                except Exception as e:
                    error_msg = str(e)
                    if "Duplicate label names found" in error_msg:
                        st.error("❌ Cannot merge files with duplicate label names")
                        st.error(error_msg)
                    else:
                        st.error(f"❌ Error merging files: {str(e)}")
                        import traceback
                        st.error(traceback.format_exc())


# ─────────────────────────────────────────────────────────────────────────────
# QC HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_polygon_bbox(points):
    """Return axis-aligned bounding box of a polygon: (x_min, y_min, x_max, y_max)."""
    xs = points[0::2]
    ys = points[1::2]
    return min(xs), min(ys), max(xs), max(ys)


def bbox_of_bbox(bbox):
    """bbox format [x, y, w, h] → (x_min, y_min, x_max, y_max)."""
    x, y, w, h = bbox
    return x, y, x + w, y + h


def aabb_overlap(a, b):
    """Check whether two AABBs (x1,y1,x2,y2) overlap."""
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    return not (ax2 <= bx1 or bx2 <= ax1 or ay2 <= by1 or by2 <= ay1)


def aabb_contains(outer, inner, tolerance=0):
    """Return True if inner AABB is fully inside outer AABB (with optional tolerance)."""
    ox1, oy1, ox2, oy2 = outer
    ix1, iy1, ix2, iy2 = inner
    return (ix1 >= ox1 - tolerance and iy1 >= oy1 - tolerance and
            ix2 <= ox2 + tolerance and iy2 <= oy2 + tolerance)


def annotation_aabb(ann):
    """Return AABB for any annotation (bbox or polygon)."""
    t = ann.get("type", "")
    if t == "bbox":
        return bbox_of_bbox(ann["bbox"])
    elif t == "polygon":
        pts = ann.get("points", [])
        if len(pts) >= 4:
            return get_polygon_bbox(pts)
    return None


def parse_qc_data(json_data):
    """
    Parse Datumaro JSON and return:
      - label_map: {id -> name}
      - name_to_id: {name -> id}
      - items: raw items list
    """
    data = json.loads(json_data) if isinstance(json_data, str) else json_data
    label_map = {}
    if "categories" in data and "label" in data["categories"]:
        for idx, lbl in enumerate(data["categories"]["label"].get("labels", [])):
            label_map[idx] = lbl.get("name", f"Label_{idx}")
    return label_map, {v: k for k, v in label_map.items()}, data.get("items", [])


# ─────────────────────────────────────────────────────────────────────────────
# CHECK 1 – Wrong annotation type for a label
# ─────────────────────────────────────────────────────────────────────────────

def qc_check1_wrong_type(items, label_map, label_name, expected_type):
    """
    Find annotations of *label_name* that are of the OPPOSITE type than
    *expected_type*.  Returns a list of dicts with frame_id, image_name,
    label_name, wrong_type and count (only items with count > 0).
    """
    wrong_type = "polygon" if expected_type == "bbox" else "bbox"
    results = []
    for frame_id, item in enumerate(items):
        image_name = item.get("id", "")
        count = 0
        for ann in item.get("annotations", []):
            lid = ann.get("label_id", -1)
            name = label_map.get(lid, f"Unknown_{lid}")
            if name == label_name and ann.get("type", "") == wrong_type:
                count += 1
        if count > 0:
            results.append({
                "frame_id": frame_id,
                "image_name": image_name,
                "label_name": label_name,
                "wrong_type": wrong_type,
                "count": count,
            })
    return results


def qc_check1_label_summary(items, label_map):
    """Return a DataFrame: label | bbox_count | polygon_count | total."""
    counts = {}
    for item in items:
        for ann in item.get("annotations", []):
            lid = ann.get("label_id", -1)
            name = label_map.get(lid, f"Unknown_{lid}")
            t = ann.get("type", "")
            if name not in counts:
                counts[name] = {"BBox": 0, "Polygon": 0}
            if t == "bbox":
                counts[name]["BBox"] += 1
            elif t == "polygon":
                counts[name]["Polygon"] += 1
    rows = [{"Label": k, "BBox": v["BBox"], "Polygon": v["Polygon"],
             "Total": v["BBox"] + v["Polygon"]} for k, v in sorted(counts.items())]
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# CHECK 2 – Nested labels inside a selected label
# ─────────────────────────────────────────────────────────────────────────────

def _point_in_polygon(px, py, flat_points):
    """Ray-casting: is (px,py) strictly inside the polygon given as flat [x,y,...] list?"""
    xs = flat_points[0::2]
    ys = flat_points[1::2]
    n = len(xs)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = xs[i], ys[i]
        xj, yj = xs[j], ys[j]
        if ((yi > py) != (yj > py)) and (px < (xj - xi) * (py - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside


def child_fully_inside_parent(child_ann, parent_ann, tolerance=5):
    """
    Return True when the child annotation is **fully contained** inside
    the parent annotation shape.

    Logic (same for Check 2 and Check 4 — both are the same concept):

    • parent is bbox    → child AABB must be fully within parent AABB.
    • parent is polygon → (a) child AABB must be fully within parent AABB
                          (b) ALL FOUR CORNERS of child AABB must pass a
                              ray-casting test against the parent polygon.

    Why all-four-corners instead of centre-only?
    Walls sit exactly on the Room polygon boundary — their centre can land
    just inside the bbox but their corners straddle the polygon edge.
    Requiring all four corners to be inside the polygon correctly excludes
    boundary-hugging walls and includes only genuine interior fixtures.
    """
    child_box = annotation_aabb(child_ann)
    if child_box is None:
        return False

    p_type = parent_ann.get("type", "")

    if p_type == "bbox":
        parent_box = annotation_aabb(parent_ann)
        if parent_box is None:
            return False
        return aabb_contains(parent_box, child_box, tolerance=tolerance)

    elif p_type == "polygon":
        pts = parent_ann.get("points", [])
        if len(pts) < 6:
            return False
        parent_box = annotation_aabb(parent_ann)
        if parent_box is None:
            return False
        # Quick AABB pre-filter
        if not aabb_contains(parent_box, child_box, tolerance=tolerance):
            return False
        # All four corners of child must be strictly inside the parent polygon
        cx1, cy1, cx2, cy2 = child_box
        corners = [(cx1, cy1), (cx2, cy1), (cx2, cy2), (cx1, cy2)]
        return all(_point_in_polygon(px, py, pts) for px, py in corners)

    return False


def _annotation_area(ann):
    """Area of an annotation's bounding box in square pixels."""
    box = annotation_aabb(ann)
    if box:
        return (box[2] - box[0]) * (box[3] - box[1])
    return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# CHECK 2 – Nested labels inside a selected label
# (same containment concept as Check 4)
# ─────────────────────────────────────────────────────────────────────────────

def qc_check2_nested_labels(items, label_map, parent_label_name, ignore_labels, tolerance=5):
    """
    For every annotation of *parent_label_name*, compute its area and report
    which other annotations are **fully contained** inside it.

    Containment rule (same as Check 4):
      • bbox parent  → child AABB fully within parent AABB
      • polygon parent → all 4 child-corners inside the parent polygon
        (this correctly excludes walls that straddle the polygon boundary)

    Returns a list of result dicts, one per parent annotation found.
    """
    results = []
    for item in items:
        anns = item.get("annotations", [])
        image_id = item.get("id", "")

        parents = [a for a in anns
                   if label_map.get(a.get("label_id", -1)) == parent_label_name]

        for p_ann in parents:
            p_id = p_ann.get("id", "N/A")
            p_area = _annotation_area(p_ann)
            p_box = annotation_aabb(p_ann)
            nested = []

            for a in anns:
                a_name = label_map.get(a.get("label_id", -1), "")
                if a_name == parent_label_name:
                    continue
                if a_name in ignore_labels:
                    continue
                if child_fully_inside_parent(a, p_ann, tolerance=tolerance):
                    a_area = _annotation_area(a)
                    nested.append({
                        "nested_label": a_name,
                        "nested_id": a.get("id", "N/A"),
                        "nested_type": a.get("type", ""),
                        "nested_area_px": round(a_area, 1),
                        "nested_area_pct": round(a_area / p_area * 100, 2) if p_area > 0 else 0.0,
                    })

            results.append({
                "image_id": image_id,
                "parent_ann_id": p_id,
                "parent_type": p_ann.get("type", ""),
                "parent_area_px": round(p_area, 1),
                "parent_bbox": (f"({p_box[0]:.1f},{p_box[1]:.1f})"
                                f" → ({p_box[2]:.1f},{p_box[3]:.1f})") if p_box else "N/A",
                "nested_count": len(nested),
                "nested_labels": ", ".join(sorted(set(n["nested_label"] for n in nested))) if nested else "—",
                "nested_details": nested,
            })
    return results


# ─────────────────────────────────────────────────────────────────────────────
# CHECK 3 – Gap between Room and surrounding labels (walls/windows/doors)
# ─────────────────────────────────────────────────────────────────────────────

def qc_check3_room_gap(items, label_map, room_label, surrounding_labels, gap_threshold=10):
    """
    For each Room polygon, check whether every side of its bounding box
    has at least one surrounding-label annotation within `gap_threshold` pixels.
    Returns issues list.
    """
    issues = []
    for item in items:
        anns = item.get("annotations", [])
        image_id = item.get("id", "")

        rooms = [a for a in anns if label_map.get(a.get("label_id", -1)) == room_label]
        surrounders = [a for a in anns if label_map.get(a.get("label_id", -1)) in surrounding_labels]

        for room in rooms:
            r_box = annotation_aabb(room)
            if r_box is None:
                continue
            rx1, ry1, rx2, ry2 = r_box
            room_id = room.get("id", "N/A")

            gaps = []
            nearby = []
            for s in surrounders:
                s_box = annotation_aabb(s)
                if s_box is None:
                    continue
                sx1, sy1, sx2, sy2 = s_box
                # Expand room box by gap_threshold and check overlap
                expanded = (rx1 - gap_threshold, ry1 - gap_threshold,
                            rx2 + gap_threshold, ry2 + gap_threshold)
                if aabb_overlap(expanded, s_box):
                    nearby.append({
                        "id": s.get("id", "N/A"),
                        "label": label_map.get(s.get("label_id", -1), ""),
                        "type": s.get("type", ""),
                        "bbox": s_box
                    })

            # Check four sides
            sides = {
                "Top": [n for n in nearby if n["bbox"][3] >= ry1 - gap_threshold and n["bbox"][1] <= ry1 + gap_threshold],
                "Bottom": [n for n in nearby if n["bbox"][1] <= ry2 + gap_threshold and n["bbox"][3] >= ry2 - gap_threshold],
                "Left": [n for n in nearby if n["bbox"][2] >= rx1 - gap_threshold and n["bbox"][0] <= rx1 + gap_threshold],
                "Right": [n for n in nearby if n["bbox"][0] <= rx2 + gap_threshold and n["bbox"][2] >= rx2 - gap_threshold],
            }

            missing_sides = [side for side, nbrs in sides.items() if len(nbrs) == 0]
            if missing_sides or len(nearby) == 0:
                issues.append({
                    "image_id": image_id,
                    "room_ann_id": room_id,
                    "room_bbox": f"({rx1:.1f},{ry1:.1f}) → ({rx2:.1f},{ry2:.1f})",
                    "nearby_surrounder_count": len(nearby),
                    "missing_sides": ", ".join(missing_sides) if missing_sides else "None (but 0 surrounders)",
                    "nearby_ids": ", ".join(str(n["id"]) for n in nearby) if nearby else "—",
                    "nearby_labels": ", ".join(sorted(set(n["label"] for n in nearby))) if nearby else "—",
                    "has_gap": len(missing_sides) > 0 or len(nearby) == 0
                })
    return issues


# ─────────────────────────────────────────────────────────────────────────────
# CHECK 4 – Annotations outside Floor plan
# (same containment concept as Check 2)
# ─────────────────────────────────────────────────────────────────────────────

def qc_check4_outside_floor_plan(items, label_map, floor_label, tolerance=5):
    """
    Same containment logic as Check 2: an annotation is "inside" a floor-plan
    bbox when its own AABB is fully contained within the floor-plan AABB.

    One image can have MULTIPLE floor-plan boxes (different building sections).
    An annotation is only flagged as "outside" if it is NOT contained in ANY
    of those floor-plan boxes.
    """
    issues = []
    for item in items:
        anns = item.get("annotations", [])
        image_id = item.get("id", "")

        floor_anns = [a for a in anns
                      if label_map.get(a.get("label_id", -1)) == floor_label]
        if not floor_anns:
            continue

        # Compute area for each floor-plan box (shown in results)
        floor_boxes = []
        for fp in floor_anns:
            fp_box = annotation_aabb(fp)
            if fp_box:
                floor_boxes.append((fp, fp_box, _annotation_area(fp)))

        if not floor_boxes:
            continue

        for a in anns:
            a_name = label_map.get(a.get("label_id", -1), "")
            if a_name == floor_label:
                continue
            a_box = annotation_aabb(a)
            if a_box is None:
                continue
            a_area = _annotation_area(a)

            # Inside if fully contained in ANY floor-plan box
            in_any = any(
                child_fully_inside_parent(a, fp_ann, tolerance=tolerance)
                for fp_ann, fp_box, fp_area in floor_boxes
            )

            if not in_any:
                # Report closest floor-plan box for context
                def dist_to_box(box):
                    fx1, fy1, fx2, fy2 = box
                    ax1, ay1, ax2, ay2 = a_box
                    return (max(0, fx1 - ax2, ax1 - fx2) +
                            max(0, fy1 - ay2, ay1 - fy2))

                closest_fp_ann, closest_fp_box, closest_fp_area = min(
                    floor_boxes, key=lambda t: dist_to_box(t[1])
                )
                issues.append({
                    "image_id": image_id,
                    "num_floor_plan_boxes": len(floor_boxes),
                    "closest_floor_area_px": round(closest_fp_area, 1),
                    "closest_floor_bbox": (
                        f"({closest_fp_box[0]:.1f},{closest_fp_box[1]:.1f})"
                        f" → ({closest_fp_box[2]:.1f},{closest_fp_box[3]:.1f})"
                    ),
                    "offending_label": a_name,
                    "offending_ann_id": a.get("id", "N/A"),
                    "offending_type": a.get("type", ""),
                    "offending_area_px": round(a_area, 1),
                    "offending_bbox": (
                        f"({a_box[0]:.1f},{a_box[1]:.1f})"
                        f" → ({a_box[2]:.1f},{a_box[3]:.1f})"
                    ),
                })
    return issues


# ─────────────────────────────────────────────────────────────────────────────
# QC PAGE
# ─────────────────────────────────────────────────────────────────────────────

def annotation_qc_page():
    st.title("🔍 Annotation QC")
    st.markdown("Quality-check your Datumaro JSON annotations across four checks.")

    uploaded_file = st.file_uploader(
        "Upload a Datumaro JSON file",
        type="json",
        key="qc_uploader"
    )

    if uploaded_file is None:
        st.info("👆 Upload a Datumaro JSON file to begin.")
        return

    try:
        json_content = uploaded_file.read().decode("utf-8")
        data = json.loads(json_content)
    except Exception as e:
        st.error(f"❌ Could not parse JSON: {e}")
        return

    label_map, name_to_id, items = parse_qc_data(json_content)
    all_label_names = sorted(label_map.values())

    if not items:
        st.error("No items found in the JSON file.")
        return

    st.success(f"✅ Loaded **{len(items)}** image(s) with **{len(all_label_names)}** label(s).")
    st.divider()

    # ── Tabs for the 4 checks ─────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "✅ Check 1 – Wrong Annotation Type",
        "🔲 Check 2 – Nested Labels",
        "🏠 Check 3 – Room Gap",
        "📐 Check 4 – Floor Plan Containment"
    ])

    # ══════════════════════════════════════════════════════════════════════════
    # CHECK 1 – Wrong annotation type for a label
    # ══════════════════════════════════════════════════════════════════════════
    with tab1:
        st.subheader("✅ Wrong Annotation Type Check")
        st.markdown(
            "Select a **label** and the **expected annotation type**. "
            "This check finds images where that label has annotations of the "
            "**opposite** type (e.g. bbox annotations when polygon is expected, "
            "or vice versa)."
        )

        col_a, col_b = st.columns(2)
        with col_a:
            sel_label = st.selectbox(
                "Select label:",
                options=all_label_names,
                key="qc1_label"
            )
        with col_b:
            sel_expected = st.selectbox(
                "Expected annotation type:",
                options=["BBox", "Polygon"],
                key="qc1_expected_type"
            )

        if st.button("▶ Run Check 1", use_container_width=True, key="run_qc1"):
            expected_type = "bbox" if sel_expected == "BBox" else "polygon"
            wrong_type = "polygon" if expected_type == "bbox" else "bbox"
            with st.spinner("Checking for wrong annotation types…"):
                results1 = qc_check1_wrong_type(items, label_map, sel_label, expected_type)

            total_wrong = sum(r["count"] for r in results1)

            c1, c2 = st.columns(2)
            c1.metric("Images with Wrong Type", len(results1))
            c2.metric(f"Total {wrong_type.title()} Annotations Found", total_wrong)
            st.divider()

            if not results1:
                st.success(f"✅ No {wrong_type} annotations found for label '{sel_label}'.")
            else:
                st.warning(
                    f"⚠️ Found **{total_wrong}** {wrong_type} annotation(s) for label "
                    f"'{sel_label}' across **{len(results1)}** image(s) "
                    f"(expected type: {expected_type})."
                )
                df1 = pd.DataFrame([{
                    "Frame ID": r["frame_id"],
                    "Image Name": r["image_name"],
                    "Label": r["label_name"],
                    "Wrong Type": r["wrong_type"],
                    "Count": r["count"],
                } for r in results1])
                st.dataframe(df1, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════════════════
    # CHECK 2 – Nested labels
    # ══════════════════════════════════════════════════════════════════════════
    with tab2:
        st.subheader("🔲 Nested Labels inside a Selected Label")
        st.markdown(
            "Select a **parent label** (e.g. *Room*). The check will find every annotation "
            "of that label and show which other labels' annotations overlap/are nested inside it. "
            "Labels in the **ignore list** will be excluded from results."
        )

        col_a, col_b = st.columns(2)
        with col_a:
            parent_label = st.selectbox(
                "Parent label to inspect:",
                options=all_label_names,
                index=all_label_names.index("Room") if "Room" in all_label_names else 0,
                key="qc2_parent"
            )
        with col_b:
            ignore_labels = st.multiselect(
                "Ignore list (exclude these from nested results):",
                options=[l for l in all_label_names if l != parent_label],
                default=[l for l in ["Floor plan", "Legend", "Sink", "Toilet", "Cooktops","Urinals", "Stairs", "Bathtub", "Shower", "Doors","Shower Door", "Sliding Door", "Single Swing Door", "Closet Door", "Double Swing Door"] if l in all_label_names and l != parent_label],
                key="qc2_ignore"
            )

        image_ids2 = ["All images"] + [item.get("id", "") for item in items]
        sel_img2 = st.selectbox("Filter by image:", image_ids2, key="qc2_img")
        check2_items = [i for i in items if i.get("id") == sel_img2] if sel_img2 != "All images" else items

        if st.button("▶ Run Check 2", use_container_width=True, key="run_qc2"):
            with st.spinner("Analysing nested labels…"):
                results2 = qc_check2_nested_labels(check2_items, label_map, parent_label, set(ignore_labels))

            has_nested = [r for r in results2 if r["nested_details"]]
            if not results2:
                st.warning(f"No '{parent_label}' annotations found.")
            elif not has_nested:
                st.info(f"No nested labels found inside any '{parent_label}' annotation.")
            else:
                df2 = pd.DataFrame([{
                    "Image": r["image_id"],
                    "Frame ID": r["parent_ann_id"],
                    "Label Type": r["parent_type"],
                    "Nested Labels": r["nested_labels"]
                } for r in has_nested])

                total_nested = sum(r["nested_count"] for r in has_nested)
                c1, c2, c3 = st.columns(3)
                c1.metric(f"Total '{parent_label}' Annotations", len(results2))
                c2.metric("With ≥1 Nested Label", len(has_nested))
                c3.metric("Total Nested Annotations", total_nested)
                st.divider()
                st.dataframe(df2, use_container_width=True, hide_index=True)

                # Drill-down expander
                st.markdown("#### 🔍 Detailed Nested Items")
                for r in has_nested:
                    with st.expander(
                        f"📦 {r['image_id']} | Frame {r['parent_ann_id']} | Parent area={r['parent_area_px']:,.0f} px²"
                        f" — {r['nested_count']} nested annotation(s): {r['nested_labels']}"
                    ):
                        nd_df = pd.DataFrame([{
                            "Nested Label": n["nested_label"],
                            "Type": n["nested_type"],
                            "Area (px²)": n["nested_area_px"],
                            "% of Parent Area": n["nested_area_pct"],
                        } for n in r["nested_details"]])
                        st.dataframe(nd_df, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════════════════
    # CHECK 3 – Room gap
    # ══════════════════════════════════════════════════════════════════════════
    with tab3:
        st.subheader("🏠 Room ↔ Surrounding Label Gap Check")
        st.markdown(
            "Each **Room** should be surrounded by walls, windows, and doors. "
            "This check verifies that at least one surrounding label is present "
            "on each of the four sides of every Room. Rooms with a missing side "
            "are flagged as having a gap."
        )

        col_r1, col_r2 = st.columns(2)
        with col_r1:
            room_label = st.selectbox(
                "Room label:",
                options=all_label_names,
                index=all_label_names.index("Room") if "Room" in all_label_names else 0,
                key="qc3_room"
            )
        with col_r2:
            surrounding_labels = st.multiselect(
                "Surrounding labels (walls / windows / doors):",
                options=[l for l in all_label_names if l != room_label],
                default=[l for l in all_label_names if any(
                    kw in l.lower() for kw in ["wall", "window", "door"]
                )],
                key="qc3_surrounding"
            )

        gap_thresh = st.slider(
            "Gap tolerance (pixels) — surrounder must be within this distance of Room edge:",
            min_value=1, max_value=100, value=15, key="qc3_thresh"
        )

        if st.button("▶ Run Check 3", use_container_width=True, key="run_qc3"):
            if not surrounding_labels:
                st.warning("Please select at least one surrounding label.")
            else:
                with st.spinner("Checking room gaps…"):
                    issues3 = qc_check3_room_gap(items, label_map, room_label, set(surrounding_labels), gap_thresh)

                room_anns_total = sum(
                    1 for item in items for a in item.get("annotations", [])
                    if label_map.get(a.get("label_id", -1)) == room_label
                )
                issues_with_gap = [i for i in issues3 if i["has_gap"]]

                c1, c2, c3 = st.columns(3)
                c1.metric("Total Room Annotations", room_anns_total)
                c2.metric("Rooms with Gaps ⚠️", len(issues_with_gap))
                c3.metric("Rooms OK ✅", room_anns_total - len(issues_with_gap))
                st.divider()

                if not issues_with_gap:
                    st.success("✅ No gap issues found! All rooms have surrounding labels on every side.")
                else:
                    st.error(f"⚠️ {len(issues_with_gap)} room(s) have gaps!")
                    df3 = pd.DataFrame([{
                        "Image": i["image_id"],
                        "Room Ann ID": i["room_ann_id"],
                        "Room BBox": i["room_bbox"],
                        "Missing Sides": i["missing_sides"],
                        "Nearby Surrounder Count": i["nearby_surrounder_count"],
                        "Nearby Labels": i["nearby_labels"],
                        "Nearby IDs": i["nearby_ids"]
                    } for i in issues_with_gap])
                    st.dataframe(df3, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════════════════
    # CHECK 4 – Floor plan containment
    # ══════════════════════════════════════════════════════════════════════════
    with tab4:
        st.subheader("📐 Floor Plan Containment Check")
        st.markdown(
            "The **Floor plan** label should act as an outer boundary that contains "
            "all other annotations. This check finds any annotation whose bounding box "
            "extends outside the Floor plan polygon's bounding box."
        )

        floor_label = st.selectbox(
            "Floor plan label:",
            options=all_label_names,
            index=all_label_names.index("Floor plan") if "Floor plan" in all_label_names else 0,
            key="qc4_floor"
        )
        tol4 = st.slider(
            "Containment tolerance (pixels) — allow this much overshoot before flagging:",
            min_value=0, max_value=50, value=5, key="qc4_tol"
        )

        image_ids4 = ["All images"] + [item.get("id", "") for item in items]
        sel_img4 = st.selectbox("Filter by image:", image_ids4, key="qc4_img")
        check4_items = [i for i in items if i.get("id") == sel_img4] if sel_img4 != "All images" else items

        if st.button("▶ Run Check 4", use_container_width=True, key="run_qc4"):
            with st.spinner("Checking floor plan containment…"):
                issues4 = qc_check4_outside_floor_plan(check4_items, label_map, floor_label, tol4)

            images_with_floor = sum(
                1 for item in check4_items
                if any(label_map.get(a.get("label_id", -1)) == floor_label for a in item.get("annotations", []))
            )

            c1, c2, c3 = st.columns(3)
            c1.metric("Images with Floor Plan", images_with_floor)
            c2.metric("Violations Found ⚠️", len(issues4))
            c3.metric("Unique Images Affected", len(set(i["image_id"] for i in issues4)))
            st.divider()

            if not issues4:
                st.success("✅ All annotations are inside the Floor plan boundary!")
            else:
                st.error(f"⚠️ {len(issues4)} annotation(s) found outside the Floor plan!")
                df4 = pd.DataFrame([{
                    "Image": i["image_id"],
                    "Floor Plan Boxes in Image": i["num_floor_plan_boxes"],
                    "Closest Floor Area (px²)": i["closest_floor_area_px"],
                    "Closest Floor BBox": i["closest_floor_bbox"],
                    "Offending Label": i["offending_label"],
                    "Offending Type": i["offending_type"],
                    "Offending Area (px²)": i["offending_area_px"],
                    "Offending BBox": i["offending_bbox"],
                } for i in issues4])

                # Group by image for clarity
                affected_images = df4["Image"].unique()
                for img in affected_images:
                    with st.expander(f"🖼️ {img} — {len(df4[df4['Image']==img])} violation(s)"):
                        st.dataframe(
                            df4[df4["Image"] == img].drop(columns=["Image"]),
                            use_container_width=True,
                            hide_index=True
                        )

                st.divider()
                st.subheader("📋 All Violations (flat table)")
                st.dataframe(df4, use_container_width=True, hide_index=True)


def main():
    st.set_page_config(page_title="Datumaro Tools", layout="wide")

    st.markdown("""
        <style>
            .main { max-width: 1200px; margin: 0 auto; }
            [data-testid="stAppViewContainer"] { padding-top: 2rem; }
        </style>
    """, unsafe_allow_html=True)

    with st.sidebar:
        st.title("🗂️ Datumaro Tools")
        st.markdown("---")
        st.markdown("**Select a Page:**")
        page = st.radio(
            "page_nav",
            [
                "📊 Annotation Converter",
                "🔷 Polygon Point Counter",
                "🔀 JSON Split & Merge",
                "🔍 Annotation QC"
            ],
            label_visibility="collapsed"
        )
        st.markdown("---")
        st.markdown("**📊 Annotation Converter**")
        st.caption("Convert Datumaro JSON to Excel with label breakdowns.")
        st.markdown("**🔷 Polygon Point Counter**")
        st.caption("Analyze per-polygon vertex (point) counts.")
        st.markdown("**🔀 JSON Split & Merge**")
        st.caption("Filter annotations by label or merge multiple JSON files.")
        st.markdown("**🔍 Annotation QC**")
        st.caption("Run 4 quality checks on your Datumaro JSON annotations.")

    if page == "📊 Annotation Converter":
        annotation_converter_page()
    elif page == "🔷 Polygon Point Counter":
        polygon_point_counter_page()
    elif page == "🔀 JSON Split & Merge":
        json_split_merge_page()
    else:
        annotation_qc_page()


if __name__ == "__main__":
    main()